from django.core.paginator import Paginator
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render

from customer.orders.models import Order

from .utils import get_sales_report_data

from io import BytesIO

from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph
from reportlab.platypus import SimpleDocTemplate
from reportlab.platypus import Spacer
from reportlab.platypus import Table
from reportlab.platypus import TableStyle
from django.utils import timezone

from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter


@staff_member_required
def sales_report(request):

    data = get_sales_report_data(request)

    orders = data["orders"]

    paginator = Paginator(orders, 10)

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    summary = data["summary"]

    context = {

        "page_obj": page_obj,

        "orders": page_obj,

        "report": data["report_type"],

        "selected_date": data["selected_date"],

        "week_from": data["week_from"],

        "week_to": data["week_to"],

        "month": data["month"],

        "year": data["year"],

        "from_date": data["from_date"],

        "to_date": data["to_date"],

        "total_orders": summary["total_orders"],

        "total_sales": summary["total_sales"],

        "subtotal": summary["subtotal"],

        "total_discount": summary["total_discount"],

        "coupon_discount": summary["coupon_discount"],

        "delivery_charge": summary["delivery_charge"],

        "tax_amount": summary["tax_amount"],

        "total_products": summary["total_products"],

        "average_order_value": summary["average_order_value"],

        "net_revenue": (
            summary["total_sales"]
            - summary["coupon_discount"]
        ),

    }

    return render( request, "reports/sales_report.html", context)


@staff_member_required
def sales_report_pdf(request):

    data = get_sales_report_data(request)

    orders = data["orders"]

    summary = data["summary"]

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20,
        leftMargin=20,
        topMargin=25,
        bottomMargin=25,
    )

    styles = getSampleStyleSheet()

    title_style = styles["Heading1"]
    title_style.alignment = TA_CENTER

    heading_style = styles["Heading3"]

    normal_style = styles["BodyText"]

    elements = []

    elements.append(
        Paragraph(
            "PocketStore Sales Report",
            title_style,
        )
    )

    elements.append(
        Spacer(
            1,
            0.25 * inch,
        )
    )

    report_info = [
        [
            "Report Type",
            data["report_type"].title(),
        ],
        [
            "Generated On",
            timezone.localtime().strftime(
                "%d-%m-%Y %I:%M %p"
            ),
        ],
        [
            "Total Orders",
            str(summary["total_orders"]),
        ],
        [
            "Total Sales",
            f"₹ {summary['total_sales']}",
        ],
        [
            "Subtotal",
            f"₹ {summary['subtotal']}",
        ],
        [
            "Coupon Discount",
            f"₹ {summary['coupon_discount']}",
        ],
        [
            "Total Discount",
            f"₹ {summary['total_discount']}",
        ],
        [
            "Delivery Charge",
            f"₹ {summary['delivery_charge']}",
        ],
        [
            "Tax",
            f"₹ {summary['tax_amount']}",
        ],
        [
            "Average Order Value",
            f"₹ {summary['average_order_value']}",
        ],
    ]

    summary_table = Table(
        report_info,
        colWidths=[
            180,
            260,
        ],
    )

    summary_table.setStyle(

        TableStyle(

            [

                (
                    "BACKGROUND",
                    (0, 0),
                    (0, -1),
                    colors.HexColor("#2563EB"),
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (0, -1),
                    colors.white,
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey,
                ),

                (
                    "BACKGROUND",
                    (1, 0),
                    (1, -1),
                    colors.whitesmoke,
                ),

                (
                    "FONTNAME",
                    (0, 0),
                    (-1, -1),
                    "Helvetica-Bold",
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),

            ]

        )

    )

    elements.append(summary_table)

    elements.append(
        Spacer(
            1,
            0.30 * inch,
        )
    )

    elements.append(
        Paragraph(
            "Order Details",
            heading_style,
        )
    )

    elements.append(
        Spacer(
            1,
            0.15 * inch,
        )
    )

    table_data = [

        [
            "#",
            "Order",
            "Date",
            "Customer",
            "Items",
            "Total",
            "Payment",
            "Status",
        ]

    ]
    for index, order in enumerate(orders, start=1):

        payment_method = "-"

        if hasattr(order, "payment") and order.payment:

            payment_method = (
                order.payment.get_payment_method_display()
            )

        table_data.append(

            [

                str(index),

                order.order_number,

                order.placed_at.strftime(
                    "%d-%m-%Y"
                ),

                order.user.get_full_name()
                or order.user.username,

                str(order.total_items),

                f"₹ {order.total_amount}",

                payment_method,

                order.get_order_status_display(),

            ]

        )

    report_table = Table(

        table_data,

        colWidths=[
            35,
            80,
            65,
            120,
            40,
            70,
            80,
            70,
        ],

        repeatRows=1,

    )

    report_table.setStyle(

        TableStyle(

            [

                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#2563EB"),
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white,
                ),

                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold",
                ),

                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER",
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.grey,
                ),

                (
                    "BACKGROUND",
                    (0, 1),
                    (-1, -1),
                    colors.beige,
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    7,
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    7,
                ),

            ]

        )

    )

    elements.append(report_table)

    document.build(elements)

    pdf = buffer.getvalue()

    buffer.close()

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="sales_report.pdf"'

    response.write(pdf)

    return response


@staff_member_required
def sales_report_excel(request):

    data = get_sales_report_data(request)

    orders = data["orders"]

    summary = data["summary"]

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Sales Report"

    title_fill = PatternFill(
        fill_type="solid",
        fgColor="2563EB",
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAFD",
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    title_font = Font(
        bold=True,
        color="FFFFFF",
        size=16,
    )

    header_font = Font(
        bold=True,
        size=11,
    )

    worksheet.merge_cells("A1:H1")

    worksheet["A1"] = "PocketStore Sales Report"

    worksheet["A1"].font = title_font

    worksheet["A1"].fill = title_fill

    worksheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.append([])

    summary_rows = [

        ["Report Type", data["report_type"].title()],

        ["Generated On", timezone.localtime().strftime("%d-%m-%Y %I:%M %p")],

        ["Total Orders", summary["total_orders"]],

        ["Total Sales", float(summary["total_sales"])],

        ["Subtotal", float(summary["subtotal"])],

        ["Coupon Discount", float(summary["coupon_discount"])],

        ["Total Discount", float(summary["total_discount"])],

        ["Delivery Charge", float(summary["delivery_charge"])],

        ["Tax Amount", float(summary["tax_amount"])],

        ["Average Order Value", float(summary["average_order_value"])],

    ]

    start_row = 3

    for row in summary_rows:

        worksheet.cell(row=start_row, column=1).value = row[0]
        worksheet.cell(row=start_row, column=2).value = row[1]

        worksheet.cell(row=start_row, column=1).font = header_font

        worksheet.cell(row=start_row, column=1).fill = header_fill

        worksheet.cell(row=start_row, column=1).border = thin_border
        worksheet.cell(row=start_row, column=2).border = thin_border

        start_row += 1

    start_row += 2

    headers = [

        "#",

        "Order Number",

        "Date",

        "Customer",

        "Items",

        "Total",

        "Payment",

        "Status",

    ]

    for column, heading in enumerate(headers, start=1):

        cell = worksheet.cell(
            row=start_row,
            column=column,
        )

        cell.value = heading

        cell.font = header_font

        cell.fill = title_fill

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        cell.border = thin_border
        current_row = start_row + 1

    for index, order in enumerate(orders, start=1):

        payment_method = "-"

        if hasattr(order, "payment") and order.payment:

            payment_method = (
                order.payment.get_payment_method_display()
            )

        worksheet.cell(
            row=current_row,
            column=1,
            value=index,
        ).border = thin_border

        worksheet.cell(
            row=current_row,
            column=2,
            value=order.order_number,
        ).border = thin_border

        worksheet.cell(
            row=current_row,
            column=3,
            value=order.placed_at.strftime("%d-%m-%Y"),
        ).border = thin_border

        worksheet.cell(
            row=current_row,
            column=4,
            value=order.user.get_full_name() or order.user.username,
        ).border = thin_border

        worksheet.cell(
            row=current_row,
            column=5,
            value=order.total_items,
        ).alignment = Alignment(horizontal="center")

        worksheet.cell(
            row=current_row,
            column=5,
        ).border = thin_border

        total_cell = worksheet.cell(
            row=current_row,
            column=6,
            value=float(order.total_amount),
        )

        total_cell.number_format = '₹#,##0.00'
        total_cell.border = thin_border

        worksheet.cell(
            row=current_row,
            column=7,
            value=payment_method,
        ).alignment = Alignment(horizontal="center")

        worksheet.cell(
            row=current_row,
            column=7,
        ).border = thin_border

        worksheet.cell(
            row=current_row,
            column=8,
            value=order.get_order_status_display(),
        ).alignment = Alignment(horizontal="center")

        worksheet.cell(
            row=current_row,
            column=8,
        ).border = thin_border

        current_row += 1

    for col in range(1, 9):

        max_length = 0

        column_letter = get_column_letter(col)

        for row in range(1, worksheet.max_row + 1):

            value = worksheet.cell(row=row, column=col).value

            if value is not None:

                max_length = max(
                   max_length,
                   len(str(value))
                )

        worksheet.column_dimensions[
            column_letter
            ].width = min(
            max_length + 4,
            35
        )

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="sales_report.xlsx"'

    workbook.save(response)

    return response